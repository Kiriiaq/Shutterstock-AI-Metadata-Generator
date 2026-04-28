"""Generate tests/TEST_PLAN.xlsx from the test inventory.

Run once after each test addition. Overwrites the xlsx.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLOCKER_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
MAJOR_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
MINOR_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

TODAY = datetime.now().strftime("%Y-%m-%d")
COLUMNS = [
    "ID",
    "Type",
    "Auto/Manuel",
    "Fonctionnalité",
    "Description",
    "Préconditions",
    "Étapes",
    "Résultat attendu",
    "Résultat obtenu",
    "Statut",
    "Criticité",
    "Lien fichier test",
    "Lien ID code (commit/bug)",
    "Date dernier test",
    "Commentaire",
]


def _write_header(ws):
    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30


def _set_widths(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _add_validations(ws, n_rows):
    """Dropdowns for Auto/Manuel, Statut, Criticité."""
    if n_rows < 2:
        return
    auto_dv = DataValidation(type="list", formula1='"Auto,Manuel"', allow_blank=False)
    statut_dv = DataValidation(type="list", formula1='"PASS,FAIL,SKIP,NON TESTÉ,XFAIL"', allow_blank=False)
    crit_dv = DataValidation(type="list", formula1='"Bloquant,Majeur,Mineur"', allow_blank=False)
    last = n_rows
    auto_dv.add(f"C2:C{last}")
    statut_dv.add(f"J2:J{last}")
    crit_dv.add(f"K2:K{last}")
    ws.add_data_validation(auto_dv)
    ws.add_data_validation(statut_dv)
    ws.add_data_validation(crit_dv)


def _color_status(ws, n_rows):
    """Color the Statut cells based on value."""
    for r in range(2, n_rows + 1):
        cell = ws.cell(row=r, column=10)  # Statut
        if cell.value == "PASS":
            cell.fill = PASS_FILL
        elif cell.value == "FAIL":
            cell.fill = FAIL_FILL
        crit_cell = ws.cell(row=r, column=11)
        if crit_cell.value == "Bloquant":
            crit_cell.fill = BLOCKER_FILL
            crit_cell.font = Font(color="FFFFFF", bold=True)
        elif crit_cell.value == "Majeur":
            crit_cell.fill = MAJOR_FILL
            crit_cell.font = Font(bold=True)
        elif crit_cell.value == "Mineur":
            crit_cell.fill = MINOR_FILL


def _add_rows(ws, rows):
    for row in rows:
        ws.append(row)


# ============================== Test inventory =============================

TESTS_UNITAIRES = [
    [
        "U-001",
        "Unitaire",
        "Auto",
        "ShutterstockParams defaults",
        "Les valeurs par défaut de ShutterstockParams sont conformes",
        "Aucune",
        "1. Instancier ShutterstockParams() ; 2. Vérifier source_folder, prefilter_enabled, model_name, min_megapixels",
        "Toutes les valeurs par défaut correspondent à la spec",
        "PASS",
        "PASS",
        "Mineur",
        "tests/test_core/test_config.py",
        "src/core/params.py",
        TODAY,
        "",
    ],
    [
        "U-002",
        "Unitaire",
        "Auto",
        "ShutterstockParams to_dict",
        "Sérialisation dict d'une instance",
        "Aucune",
        "1. ShutterstockParams(source_folder='/x') ; 2. .to_dict() ; 3. Vérifier le dict renvoyé",
        "Dict contient source_folder='/x'",
        "PASS",
        "PASS",
        "Mineur",
        "tests/test_core/test_config.py",
        "src/core/params.py",
        TODAY,
        "",
    ],
    [
        "U-003",
        "Unitaire",
        "Auto",
        "ShutterstockParams from_dict",
        "Reconstruction depuis dict (clés invalides ignorées)",
        "Aucune",
        "1. from_dict({'source_folder': '/x', 'invalid_key': 'val'}) ; 2. Vérifier source_folder + absence d'invalid_key",
        "source_folder='/x', pas d'invalid_key",
        "PASS",
        "PASS",
        "Mineur",
        "tests/test_core/test_config.py",
        "src/core/params.py",
        TODAY,
        "",
    ],
    [
        "U-004",
        "Unitaire",
        "Auto",
        "PARAMS_META cohérence",
        "Le dict PARAMS_META est défini et non vide, source_folder existe",
        "Aucune",
        "1. Importer PARAMS_META ; 2. Vérifier non-None et len > 0 ; 3. Vérifier 'source_folder' avec label/category",
        "PARAMS_META['source_folder'].label == 'Dossier source', category == 'essential'",
        "PASS",
        "PASS",
        "Mineur",
        "tests/test_core/test_config.py",
        "src/core/params.py",
        TODAY,
        "",
    ],
    [
        "U-005",
        "Unitaire",
        "Auto",
        "validate_image_dimensions OK",
        "Validation passe pour 3000x2000 @ 4MP",
        "Aucune",
        "1. validate_image_dimensions(3000, 2000, min_megapixels=4.0)",
        "(True, None)",
        "PASS",
        "PASS",
        "Mineur",
        "tests/test_utils/test_validators.py",
        "src/utils/validators.py",
        TODAY,
        "",
    ],
    [
        "U-006",
        "Unitaire",
        "Auto",
        "validate_image_dimensions reject",
        "Rejet d'une image sub-megapixel",
        "Aucune",
        "1. validate_image_dimensions(1000, 1000, min_megapixels=4.0)",
        "(False, message contenant 'too low')",
        "PASS",
        "PASS",
        "Mineur",
        "tests/test_utils/test_validators.py",
        "src/utils/validators.py",
        TODAY,
        "",
    ],
    [
        "U-007",
        "Unitaire",
        "Auto",
        "validate_metadata_completeness 100%",
        "Score 100 quand title + description + 7 keywords + 2 categories",
        "Aucune",
        "1. Appeler avec full set ; 2. Vérifier completeness_score == 100",
        "completeness_score == 100",
        "PASS",
        "PASS",
        "Mineur",
        "tests/test_utils/test_validators.py",
        "src/utils/validators.py",
        TODAY,
        "",
    ],
    [
        "U-008",
        "Unitaire",
        "Auto",
        "validate_metadata_completeness missing title",
        "Échec si title None",
        "Aucune",
        "1. Appeler sans title ; 2. Vérifier is_valid==False et erreur 'title'",
        "is_valid False, errors mentionne title",
        "PASS",
        "PASS",
        "Mineur",
        "tests/test_utils/test_validators.py",
        "src/utils/validators.py",
        TODAY,
        "",
    ],
    [
        "U-009",
        "Unitaire",
        "Auto",
        "validate_metadata_completeness insufficient keywords",
        "Échec si < 7 keywords",
        "Aucune",
        "1. Appeler avec 2 keywords ; 2. is_valid False, erreur 'keyword'",
        "is_valid False, errors mentionne keyword",
        "PASS",
        "PASS",
        "Mineur",
        "tests/test_utils/test_validators.py",
        "src/utils/validators.py",
        TODAY,
        "",
    ],
]

TESTS_SMOKE = [
    [
        "S-001",
        "Smoke",
        "Auto",
        "Database CRUD",
        "set_setting/get_setting + log_action/get_audit_logs cycle",
        "Aucune",
        "1. Database tmp ; 2. set/get setting ; 3. log_action ; 4. get_audit_logs ; 5. close",
        "log_id > 0, len(logs)==1, file_path correct",
        "PASS",
        "PASS",
        "Bloquant",
        "tests/smoke/test_smoke.py::test_database_crud",
        "src/modules/storage/database.py",
        TODAY,
        "Garde-fou Phase E",
    ],
    [
        "S-002",
        "Smoke",
        "Auto",
        "Database batch lifecycle",
        "create_batch + update_batch_progress + complete_batch",
        "Aucune",
        "1. create_batch ; 2. update progress ; 3. complete ; 4. get_statistics",
        "stats['total_batches'] == 1",
        "PASS",
        "PASS",
        "Bloquant",
        "tests/smoke/test_smoke.py::test_database_batch_lifecycle",
        "src/modules/storage/database.py / B-4 / B-5",
        TODAY,
        "",
    ],
    [
        "S-003",
        "Smoke",
        "Auto",
        "Database set_file_flags partial update",
        "Met à jour has_metadata/has_ai_analysis sans recopier hash/size",
        "Aucune",
        "1. set_file_flags(path, has_ai_analysis=True) ; 2. get_file_status ; 3. set_file_flags(has_metadata=True) ; 4. get_file_status",
        "Les flags persistent et s'accumulent",
        "PASS",
        "PASS",
        "Bloquant",
        "tests/smoke/test_smoke.py::test_database_set_file_flags",
        "src/modules/storage/database.py / B-3",
        TODAY,
        "Méthode introduite par l'audit",
    ],
    [
        "S-004",
        "Smoke",
        "Auto",
        "IPTCEngine list_templates",
        "Liste des templates par défaut accessible",
        "Aucune",
        "1. Instancier IPTCEngine ; 2. list_templates() ; 3. Vérifier list",
        "Renvoie list",
        "PASS",
        "PASS",
        "Mineur",
        "tests/smoke/test_smoke.py::test_iptc_engine_templates",
        "src/modules/engines/iptc_engine.py",
        TODAY,
        "",
    ],
    [
        "S-005",
        "Smoke",
        "Auto",
        "IPTCFields scalars roundtrip",
        "to_dict/from_dict pour champs scalaires (headline, caption, country_code…)",
        "Aucune",
        "1. IPTCFields(...) ; 2. .to_dict() ; 3. .from_dict(d) ; 4. Vérifier valeurs",
        "Champs scalaires identiques après roundtrip",
        "PASS",
        "PASS",
        "Majeur",
        "tests/smoke/test_smoke.py::test_iptc_fields_roundtrip_scalars",
        "src/modules/models/metadata_models.py",
        TODAY,
        "",
    ],
    [
        "S-006",
        "Smoke",
        "Auto",
        "IPTCFields lists roundtrip (B-16 fix)",
        "to_dict/from_dict pour keywords + supplemental_categories",
        "Aucune",
        "1. IPTCFields(keywords=[...], supplemental_categories=[...]) ; 2. roundtrip ; 3. Vérifier listes",
        "Listes identiques après roundtrip",
        "PASS",
        "PASS",
        "Bloquant",
        "tests/smoke/test_smoke.py::test_iptc_fields_roundtrip_lists",
        "src/modules/models/metadata_models.py / B-16",
        TODAY,
        "Régression guard pour le bug B-16 corrigé",
    ],
    [
        "S-007",
        "Smoke",
        "Auto",
        "collect_image_files flat + recursive",
        "Détection .jpg/.png case-insensitive, recursion",
        "Aucune",
        "1. Créer .jpg/.PNG/.txt + sub/.tiff ; 2. recursive=False ; 3. recursive=True",
        "flat={'a.jpg','b.png'}; deep={'a.jpg','b.png','c.tiff'}",
        "PASS",
        "PASS",
        "Majeur",
        "tests/smoke/test_smoke.py::test_collect_image_files",
        "src/modules/workers/worker_pool.py",
        TODAY,
        "",
    ],
    [
        "S-008",
        "Smoke",
        "Auto",
        "clean_keywords_advanced",
        "Stopwords retirés, doublons dédupliqués, len ≥ 2",
        "Aucune",
        "1. clean_keywords_advanced(['mountain','the','Mountain','x','  trees  ','photo'])",
        "Pas de stopword ; 'mountain' une seule fois ; 'trees' présent",
        "PASS",
        "PASS",
        "Mineur",
        "tests/smoke/test_smoke.py::test_clean_keywords_basic",
        "src/modules/workers/worker_pool.py",
        TODAY,
        "",
    ],
    [
        "S-009",
        "Smoke",
        "Auto",
        "WorkerPool start/stop",
        "Lifecycle pool propre",
        "Aucune",
        "1. WorkerPool(2) ; 2. start ; 3. running True ; 4. stop ; 5. running False",
        "_running flip True puis False",
        "PASS",
        "PASS",
        "Majeur",
        "tests/smoke/test_smoke.py::test_worker_pool_start_stop",
        "src/modules/workers/worker_pool.py",
        TODAY,
        "",
    ],
    [
        "S-010",
        "Smoke",
        "Auto",
        "WorkerPool execute handler",
        "Exécution réelle d'un handler enregistré",
        "Aucune",
        "1. register_handler('noop',…) ; 2. submit_task ; 3. process_queue ; 4. Vérifier completed_tasks=1",
        "completed=1, failed=0",
        "PASS",
        "PASS",
        "Majeur",
        "tests/smoke/test_smoke.py::test_worker_pool_executes_handler",
        "src/modules/workers/worker_pool.py",
        TODAY,
        "",
    ],
    [
        "S-011",
        "Smoke",
        "Auto",
        "OllamaStatus enum",
        "Valeurs ONLINE/OFFLINE/BUSY conformes",
        "Aucune",
        "1. Vérifier OllamaStatus.ONLINE.value=='online' etc.",
        "online/offline/busy",
        "PASS",
        "PASS",
        "Mineur",
        "tests/smoke/test_smoke.py::test_ollama_status_enum",
        "src/modules/ai/ollama_client.py",
        TODAY,
        "",
    ],
    [
        "S-012",
        "Smoke",
        "Auto",
        "src package imports end-to-end",
        "15 modules clés s'importent sans erreur (catch des chaînes circulaires)",
        "Aucune",
        "1. importlib.import_module sur 15 chemins src.*",
        "Aucune ImportError",
        "PASS",
        "PASS",
        "Bloquant",
        "tests/smoke/test_smoke.py::test_src_package_imports",
        "src/* / B-17",
        TODAY,
        "Critique car B-17 a montré qu'on pouvait avoir des modules non chargeables",
    ],
    [
        "S-013",
        "Smoke",
        "Auto",
        "ShutterstockAIv2 graceful sans ExifTool",
        "La façade s'instancie même quand ExifTool absent",
        "Aucune",
        "1. Mock MetadataReader pour lever ExifToolNotFoundError ; 2. Instancier ShutterstockAIv2 ; 3. Vérifier exiftool_available=False, reader/writer None",
        "Construction OK, attributs None",
        "PASS",
        "PASS",
        "Majeur",
        "tests/smoke/test_smoke.py::test_shutterstock_ai_v2_instantiates",
        "src/modules/integration.py",
        TODAY,
        "",
    ],
]

TESTS_IHM = [
    [
        "I-001",
        "IHM",
        "Auto",
        "App full lifecycle headless",
        "App() construit, expose 6 onglets + title, on_closing() ferme proprement",
        "Tk root disponible (Windows desktop ou DISPLAY exporté). Ollama mocké pour éviter timeout 5s.",
        "1. monkeypatch requests.get/post ; 2. main.App() ; 3. update_idletasks ; 4. Vérifier title ; 5. Vérifier 6 onglets ; 6. on_closing()",
        "title contient 'ShutterstockAnalyzer v2.0.0' et 'AI Metadata Generator', 6 onglets présents, fermeture sans exception",
        "PASS",
        "PASS",
        "Bloquant",
        "tests/ui/test_app_smoke.py::test_app_full_lifecycle",
        "main.py / B-1, B-7, B-17, B-18",
        TODAY,
        "Smoke UI principal — gating Phase G",
    ],
    [
        "I-002",
        "IHM",
        "Manuel",
        "AI Control — Check Connection",
        "Bouton 'Check Connection' déclenche un essai HTTP vers Ollama et MAJ l'indicateur",
        "App lancée, Ollama démarré (ou non) à localhost:11434",
        "1. Aller onglet AI Control ; 2. Cliquer 'Check Connection' ; 3. Observer indicateur",
        "Indicateur passe à Online/Offline selon serveur ; bouton se réactive après le check",
        "",
        "NON TESTÉ",
        "Majeur",
        "(manuel)",
        "src/ui/pages/ai_control_page.py",
        "",
        "",
    ],
    [
        "I-003",
        "IHM",
        "Manuel",
        "Scan Images — Browse + Scan recursive",
        "Sélection dossier puis scan récursif détecte les images supportées",
        "App lancée, dossier de test prêt (avec sous-dossiers contenant .jpg/.png)",
        "1. Onglet Scan ; 2. Browse dossier ; 3. Cocher Recursive ; 4. Scan ; 5. Observer la liste",
        "Liste contient toutes les images ; count = total ; preview au clic",
        "",
        "NON TESTÉ",
        "Majeur",
        "(manuel)",
        "src/ui/pages/scan_page.py",
        "",
        "",
    ],
    [
        "I-004",
        "IHM",
        "Manuel",
        "Metadata Editor — Read/Write",
        "Lit puis écrit les IPTC sur une image (avec ExifTool installé)",
        "ExifTool dans PATH, image .jpg de test",
        "1. Onglet Metadata Editor ; 2. Select Folder ; 3. Cliquer un fichier ; 4. Modifier headline ; 5. 'Write to File'",
        "Read affiche IPTC actuels ; après Write, succès messagebox + relecture cohérente",
        "",
        "NON TESTÉ",
        "Bloquant",
        "(manuel)",
        "src/ui/pages/write_page.py / B-1",
        "",
        "Tab cassé avant audit (B-1)",
    ],
    [
        "I-005",
        "IHM",
        "Manuel",
        "Audit Log — filtre + double-clic",
        "Filtres action/date/status fonctionnent ; double-clic affiche les détails",
        "App lancée avec quelques entrées d'audit en DB (lancer un read/write avant)",
        "1. Onglet Audit Log ; 2. Changer filtres ; 3. Vérifier table ; 4. Double-clic une ligne",
        "Table filtrée correctement ; details_text affiche timestamp/action/status/duration/batch_id/error",
        "",
        "NON TESTÉ",
        "Majeur",
        "(manuel)",
        "src/ui/pages/audit_page.py",
        "",
        "",
    ],
    [
        "I-006",
        "IHM",
        "Manuel",
        "Settings — save then reopen",
        "Sauvegarder une valeur, fermer, rouvrir : la valeur persiste en DB",
        "App lancée",
        "1. Onglet Settings ; 2. Modifier Default Copyright ; 3. 'Save Settings' ; 4. Fermer App ; 5. Rouvrir ; 6. Vérifier la valeur",
        "La valeur saisie est rechargée à l'ouverture suivante",
        "",
        "NON TESTÉ",
        "Majeur",
        "(manuel)",
        "src/ui/pages/settings_page.py",
        "",
        "",
    ],
    [
        "I-007",
        "IHM",
        "Manuel",
        "Stubs désactivés sont visibles",
        "Les boutons coming-soon sont grisés et labellisés clairement",
        "App lancée",
        "1. Settings → 'Test (coming soon)' (FTPS) ; 2. Settings → 'Create New Template (coming soon)' ; 3. Editor → 'Write to All Files (coming soon)'",
        "Les 3 boutons sont state=disabled, label '(coming soon)' visible",
        "",
        "NON TESTÉ",
        "Mineur",
        "(manuel)",
        "src/ui/pages/settings_page.py + write_page.py / B-11, B-12",
        "",
        "Vérification UX honnêteté",
    ],
]

TESTS_REGRESSION = [
    [
        "R-001",
        "Régression",
        "Auto",
        "B-16 IPTCFields list fields",
        "Cf. S-006",
        "—",
        "Cf. S-006",
        "Cf. S-006",
        "PASS",
        "PASS",
        "Bloquant",
        "tests/smoke/test_smoke.py::test_iptc_fields_roundtrip_lists",
        "B-16",
        TODAY,
        "Sentinelle xfail-strict promue régression guard",
    ],
    [
        "R-002",
        "Régression",
        "Auto",
        "B-2 add_audit_log API",
        "Cf. S-001 + S-013",
        "—",
        "Toute opération AI déclenche log_action sans AttributeError",
        "Pas d'AttributeError au runtime",
        "PASS",
        "PASS",
        "Bloquant",
        "tests/smoke/test_smoke.py::test_database_crud",
        "B-2",
        TODAY,
        "Surface couverte indirectement",
    ],
    [
        "R-003",
        "Régression",
        "Auto",
        "B-3 set_file_flags ne casse pas update_file_status",
        "Le UPDATE partiel ne crée pas un état cassé pour les futures full updates",
        "—",
        "Cf. S-003",
        "Flags persistent + row INSERT-OR-IGNORE OK",
        "PASS",
        "PASS",
        "Bloquant",
        "tests/smoke/test_smoke.py::test_database_set_file_flags",
        "B-3",
        TODAY,
        "",
    ],
    [
        "R-004",
        "Régression",
        "Auto",
        "B-17 src package imports",
        "Aucun module n'a un import broken au top-level",
        "—",
        "Cf. S-012",
        "Tous les modules s'importent",
        "PASS",
        "PASS",
        "Bloquant",
        "tests/smoke/test_smoke.py::test_src_package_imports",
        "B-17",
        TODAY,
        "B-17 a appris la leçon — couverture par import bulk",
    ],
]

TESTS_EXE = [
    [
        "E-001",
        "EXE",
        "Manuel",
        "Build debug profile",
        "python build.py debug produit l'EXE attendu",
        "PyInstaller installé, working tree clean",
        "1. python build.py debug ; 2. Vérifier dist/ShutterstockAnalyzer-debug.exe ; 3. Taille raisonnable",
        "Build OK ; EXE ~24 MB ; --debug=imports + --console",
        "PASS",
        "PASS",
        "Majeur",
        "build.py",
        "BUILD_REPORT.md",
        TODAY,
        "Run du Phase G : 24.4 MB",
    ],
    [
        "E-002",
        "EXE",
        "Manuel",
        "Build release profile",
        "python build.py release produit l'EXE windowed",
        "PyInstaller installé",
        "1. python build.py release ; 2. Vérifier dist/ShutterstockAnalyzer.exe",
        "Build OK ; --windowed --noconsole --noupx",
        "PASS",
        "PASS",
        "Bloquant",
        "build.py",
        "BUILD_REPORT.md",
        TODAY,
        "24.4 MB",
    ],
    [
        "E-003",
        "EXE",
        "Manuel",
        "EXE release démarre sans console",
        "Lancer ShutterstockAnalyzer.exe en double-clic ; aucune console n'apparaît",
        "Windows desktop",
        "1. Double-clic dist/ShutterstockAnalyzer.exe ; 2. Observer 5 secondes",
        "Splash brief puis fenêtre principale, AUCUNE console visible",
        "",
        "NON TESTÉ",
        "Bloquant",
        "(manuel)",
        "main.py / B-7",
        "",
        "",
    ],
    [
        "E-004",
        "EXE",
        "Manuel",
        "Icône fenêtre + barre des tâches",
        "Icône custom apparaît en sup. gauche et dans la barre des tâches",
        "ShutterstockAnalyzer.exe lancé",
        "1. Observer le coin sup. gauche de la fenêtre ; 2. Observer la barre des tâches Windows",
        "Icône icone.ico (pas l'icône Python par défaut)",
        "",
        "NON TESTÉ",
        "Majeur",
        "(manuel)",
        "main.py / B-7",
        "",
        "AppUserModelID + iconbitmap branchés",
    ],
    [
        "E-005",
        "EXE",
        "Manuel",
        "Title fenêtre conforme",
        "Titre affiché : 'ShutterstockAnalyzer v2.0.0 - AI Metadata Generator for Stock Photography'",
        "App lancée",
        "1. Observer la barre de titre",
        "Format exact respecté",
        "",
        "NON TESTÉ",
        "Majeur",
        "(manuel)",
        "main.py",
        "",
        "Acceptance check Phase G",
    ],
    [
        "E-006",
        "EXE",
        "Auto",
        "Smoke launch via subprocess",
        "EXE lancé en background reste vivant > 6s (mainloop atteint sans crash)",
        "PyInstaller build complet",
        "1. subprocess.Popen(EXE, timeout=6) ; 2. Si timeout = vivant = OK",
        "Stay alive past timeout (6s)",
        "PASS",
        "PASS",
        "Bloquant",
        "build.py::_smoke",
        "BUILD_REPORT.md",
        TODAY,
        "Mode debug: 3278 lignes import trace, mode release: 0 byte (no console)",
    ],
]

ALL_SHEETS = {
    "Tests Unitaires": TESTS_UNITAIRES,
    "Tests Smoke": TESTS_SMOKE,
    "Tests IHM": TESTS_IHM,
    "Tests Régression": TESTS_REGRESSION,
    "Tests EXE": TESTS_EXE,
}


def _add_sheet(wb: Workbook, name: str, rows: list) -> None:
    ws = wb.create_sheet(title=name)
    _write_header(ws)
    _add_rows(ws, rows)
    _set_widths(ws, [10, 12, 11, 28, 38, 22, 50, 38, 30, 12, 11, 50, 30, 14, 30])

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
    _add_validations(ws, ws.max_row)
    _color_status(ws, ws.max_row)
    ws.freeze_panes = "A2"


def _add_synthese(wb: Workbook) -> None:
    ws = wb.create_sheet(title="Synthèse")
    ws["A1"] = "Synthèse — TEST_PLAN audit/20260428"
    ws["A1"].font = Font(bold=True, size=14)

    sheet_names = list(ALL_SHEETS.keys())
    headers = ["Feuille", "Total", "PASS", "FAIL", "SKIP", "NON TESTÉ", "XFAIL", "% Auto"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    row = 4
    totals = [0] * 6
    n_auto = 0
    n_total = 0
    for sname in sheet_names:
        rows = ALL_SHEETS[sname]
        total = len(rows)
        statuses = [r[9] for r in rows]
        modes = [r[2] for r in rows]
        passes = sum(1 for s in statuses if s == "PASS")
        fails = sum(1 for s in statuses if s == "FAIL")
        skips = sum(1 for s in statuses if s == "SKIP")
        non = sum(1 for s in statuses if s == "NON TESTÉ")
        xfail = sum(1 for s in statuses if s == "XFAIL")
        auto = sum(1 for m in modes if m == "Auto")
        pct_auto = (auto / total * 100) if total else 0
        ws.cell(row=row, column=1, value=sname)
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=3, value=passes).fill = PASS_FILL if passes else PatternFill()
        ws.cell(row=row, column=4, value=fails).fill = FAIL_FILL if fails else PatternFill()
        ws.cell(row=row, column=5, value=skips)
        ws.cell(row=row, column=6, value=non)
        ws.cell(row=row, column=7, value=xfail)
        ws.cell(row=row, column=8, value=f"{pct_auto:.0f}%")
        totals[0] += total
        totals[1] += passes
        totals[2] += fails
        totals[3] += skips
        totals[4] += non
        totals[5] += xfail
        n_auto += auto
        n_total += total
        row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=totals[0]).font = Font(bold=True)
    ws.cell(row=row, column=3, value=totals[1]).font = Font(bold=True)
    ws.cell(row=row, column=4, value=totals[2]).font = Font(bold=True)
    ws.cell(row=row, column=5, value=totals[3]).font = Font(bold=True)
    ws.cell(row=row, column=6, value=totals[4]).font = Font(bold=True)
    ws.cell(row=row, column=7, value=totals[5]).font = Font(bold=True)
    ws.cell(row=row, column=8, value=f"{(n_auto / n_total * 100):.0f}%" if n_total else "-").font = Font(bold=True)

    ws["A" + str(row + 3)] = "Légende criticité :"
    ws["A" + str(row + 4)] = "Bloquant"
    ws["A" + str(row + 4)].fill = BLOCKER_FILL
    ws["A" + str(row + 4)].font = Font(color="FFFFFF", bold=True)
    ws["A" + str(row + 5)] = "Majeur"
    ws["A" + str(row + 5)].fill = MAJOR_FILL
    ws["A" + str(row + 5)].font = Font(bold=True)
    ws["A" + str(row + 6)] = "Mineur"
    ws["A" + str(row + 6)].fill = MINOR_FILL

    ws["A" + str(row + 8)] = "Workflow"
    ws["A" + str(row + 8)].font = Font(bold=True)
    ws["A" + str(row + 9)] = (
        "1. Run pytest tests/ -q ; reporter PASS/FAIL dans les feuilles Tests Unitaires/Smoke/Régression."
    )
    ws["A" + str(row + 10)] = "2. Build : python build.py all ; reporter dans Tests EXE."
    ws["A" + str(row + 11)] = "3. Tests IHM : ouvrir CHECKLIST_IHM.html, exécuter les checks manuels, reporter ici."

    _set_widths(ws, [22, 10, 10, 10, 10, 14, 10, 10])


def main() -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in ALL_SHEETS.items():
        _add_sheet(wb, name, rows)
    _add_synthese(wb)

    out = Path(__file__).parent / "TEST_PLAN.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    path = main()
    print(f"Generated {path}")
