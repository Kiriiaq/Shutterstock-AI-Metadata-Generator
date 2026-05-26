"""Generate test/matrice_tests.xlsx — qualification test matrix.

Two sheets:
- ``Tests``: one row per test case, with status dropdowns.
- ``Synthèse``: counts + OK/NOK/NA rates per category (formulas).

Run once from the repo root::

    python test/scripts/_make_matrix.py
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parents[1] / "matrice_tests.xlsx"

HEADER = [
    "ID_Test",
    "Catégorie",
    "Exigence",
    "Fonctionnalité",
    "Description",
    "Pré-requis",
    "Données entrée",
    "Résultat attendu",
    "Résultat obtenu",
    "Statut",
    "Sévérité",
    "Testeur",
    "Date",
    "Commentaires",
]


# Each row: (id, cat, req, feat, desc, prereq, input, expected, severity)
TESTS = [
    # --- A. IHM (manual via validation_ihm.html) ---
    ("T-001", "IHM", "REQ-IHM-01", "Topbar / Theme",
     "Toggle theme 3× (light→dark→system→light)",
     "App lancée", "3 clics sur 🌙/☀/⚙",
     "La fenêtre garde sa taille, palette change, aucune erreur log", "mineure"),
    ("T-002", "IHM", "REQ-IHM-02", "Topbar / Aide",
     "Bouton Aide → modal raccourcis", "App lancée", "Clic ? ou Ctrl+?",
     "Modal listant tous les raccourcis, fermable par Esc", "mineure"),
    ("T-003", "IHM", "REQ-IHM-03", "Sources / Scanner",
     "Scanner un dossier de 10 images JPEG valides",
     "Dossier test préparé", "Sélectionner dossier puis cliquer Scanner",
     "DataTable affiche 10 lignes, compteur = « 10 fichiers : 10 »",
     "majeure"),
    ("T-004", "IHM", "REQ-IHM-04", "Sources / + Fichiers",
     "Ajouter 5 fichiers via le bouton + Fichiers…",
     "Liste déjà peuplée", "Multi-sélection 5 jpg dans le dialogue",
     "Ajout sans doublons, compteur incrémenté", "mineure"),
    ("T-005", "IHM", "REQ-IHM-05", "Sources / Supprimer",
     "Supprimer une sélection via clavier (Suppr) ou bouton",
     "1 ligne sélectionnée", "Touche Suppr",
     "Ligne retirée, compteur −1, focus voisin", "mineure"),
    ("T-006", "IHM", "REQ-IHM-06", "Sources / Vider",
     "Bouton Vider — action destructive avec confirmation",
     "Liste non vide", "Clic Vider → Confirmer",
     "Liste vide après confirmation, refus si Annuler", "mineure"),
    ("T-007", "IHM", "REQ-IHM-07", "Sources / Tout-Aucun",
     "Boutons Tout / Aucun pour la sélection",
     "Liste non vide", "Clic Tout puis clic Aucun",
     "Sélection complète puis vidée", "mineure"),
    ("T-008", "IHM", "REQ-IHM-08", "Éditeur / Collapse",
     "Toggle ▼/▶ masque/affiche le corps", "App lancée", "Clic flèche",
     "Corps masqué/visible, état conservé pendant la session", "mineure"),
    ("T-009", "IHM", "REQ-IHM-09", "Éditeur / Lire",
     "Double-clic ligne → 5 champs remplis depuis IPTC",
     "Image scannée avec IPTC", "Double-clic ligne Sources",
     "Titre/Description/Mots-clés/Auteur/Copyright remplis", "majeure"),
    ("T-010", "IHM", "REQ-IHM-10", "Éditeur / Écrire",
     "Modifier titre + cliquer Écrire", "Image scannée", "Modifier puis clic Écrire",
     "Toast succès, IPTC du fichier modifié sur disque (vérif avec exiftool)",
     "majeure"),
    ("T-011", "IHM", "REQ-IHM-11", "Éditeur / Rapport expert sans sélection",
     "Cliquer Rapport expert sans avoir sélectionné de fichier",
     "App lancée", "Clic Rapport expert…",
     "Toast warning « Sélectionnez d'abord une image »", "majeure"),
    ("T-012", "IHM", "REQ-IHM-12", "Éditeur / Rapport expert avec sélection",
     "Cliquer Rapport expert après double-clic sur ligne",
     "Image scannée + double-clic", "Clic Rapport expert…",
     "Modal s'ouvre, mode « Rapide (sans IA) », 8 sections affichées",
     "critique"),
    ("T-013", "IHM", "REQ-IHM-13", "Analyse / Démarrer désactivé",
     "Démarrer grisé tant que sélection vide", "Liste vide ou rien sélectionné",
     "Vérifier visuel + tentative clic", "Bouton désactivé, clic sans effet",
     "majeure"),
    ("T-014", "IHM", "REQ-IHM-14", "Analyse / Arrêter via Esc",
     "Esc annule le batch en cours", "Batch en cours", "Appui Esc",
     "Toast « Annulation demandée… », batch s'arrête à l'image suivante",
     "majeure"),
    ("T-015", "IHM", "REQ-IHM-15", "Modaux / Ouverture-fermeture",
     "Ouvrir/fermer 5 modaux (settings, audit, ai_control, validate, expert_report)",
     "App lancée", "Détail… × 5 + Esc à chaque fois",
     "Chaque modal s'ouvre puis se ferme proprement, fenêtre principale OK",
     "majeure"),

    # --- B. Paramètres ---
    ("T-020", "Paramètres", "REQ-PARAM-01", "Mode sans IA (défaut)",
     "Rapport expert avec checkbox IA décochée (défaut)",
     "Aucun Ollama requis", "Ouvrir Rapport expert sur une image",
     "Source = « heuristic », aucune trace d'appel HTTP Ollama dans les logs",
     "critique"),
    ("T-021", "Paramètres", "REQ-PARAM-02", "Use_AI=True sans Ollama",
     "Cocher IA + Régénérer alors qu'Ollama est absent",
     "Ollama service stoppé", "Cocher case puis Régénérer",
     "Fallback heuristique transparent, pas de crash, mode reste « heuristic »",
     "majeure"),
    ("T-022", "Paramètres", "REQ-PARAM-03", "Use_AI=True avec Ollama",
     "Cocher IA + Régénérer avec Ollama démarré",
     "Ollama serve + modèle llama3.2-vision installé",
     "Cocher case puis Régénérer",
     "Source = « hybrid », scores enrichis, titres potentiellement reformulés",
     "majeure"),

    # --- C. Entrées (fichiers) ---
    ("T-030", "Entrées", "REQ-IN-01", "JPEG nominal",
     "Image 12 MP, sRGB, ~3 Mo, IPTC complet",
     "input_nominal.jpg dans test/inputs/",
     "Analyse + Rapport expert",
     "Rapport complet, scores commercial≥6 / technique≥8 / SEO≥5, 0 risque blocker",
     "critique"),
    ("T-031", "Entrées", "REQ-IN-02", "Fichier vide",
     "Fichier 0 octet renommé .jpg",
     "input_vide.jpg dans test/inputs/", "Tenter scan / analyse",
     "Filtre rejette proprement (Invalid), log erreur, pas de crash UI",
     "majeure"),
    ("T-032", "Entrées", "REQ-IN-03", "Sous-résolution",
     "Image 1600×1250 (2 MP) — sous le minimum 4 MP",
     "input_low_mp.jpg", "Rapport expert",
     "RejectionRisk severity=blocker « < 4 MP », score technique ≤7",
     "majeure"),
    ("T-033", "Entrées", "REQ-IN-04", "Fichier volumineux",
     "JPEG > 50 Mo (au-dessus du plafond Shutterstock)",
     "input_volumineux.jpg (~52 Mo)", "Rapport expert",
     "Warnings Adobe + Shutterstock (poids), pas bloqué, scores OK",
     "majeure"),
    ("T-034", "Entrées", "REQ-IN-05", "Mauvais format",
     "PNG renommé / vrai PNG", "input_mauvais_format.png",
     "Rapport expert", "Warning « format JPEG recommandé » sur les deux plateformes",
     "mineure"),
    ("T-035", "Entrées", "REQ-IN-06", "Espace CMYK",
     "JPEG en mode CMYK", "input_cmyk.jpg", "Rapport expert",
     "Warning « espace CMJN détecté (sRGB attendu) », score tech −1",
     "mineure"),
    ("T-036", "Entrées", "REQ-IN-07", "UTF-8 / accents / symboles I&C",
     "IPTC contient « Çà éü ° ± μ Ω 🎨 »",
     "input_utf8.jpg avec IPTC riche", "Lire IPTC dans l'éditeur",
     "Tous caractères préservés à l'écriture, encoding utf-8 OK",
     "majeure"),
    ("T-037", "Entrées", "REQ-IN-08", "Fichier corrompu",
     "Header JPEG tronqué (premiers 200 octets seulement)",
     "input_corrompu.jpg", "Scanner le dossier",
     "Erreur PIL contenue, ligne marquée Invalid, app stable",
     "majeure"),
    ("T-038", "Entrées", "REQ-IN-09", "Chemin avec espaces + accents",
     "Image placée dans « Dossier été/photo n°1.jpg »",
     "Sous-dossier `Dossier été/`", "Scanner + analyser",
     "Chemins gérés correctement (pas de UnicodeError, pas de quoting cassé)",
     "mineure"),

    # --- D. Sorties ---
    ("T-040", "Sorties", "REQ-OUT-01", "CSV Adobe — colonnes",
     "Export CSV Adobe après un rapport expert",
     "Rapport expert calculé", "Bouton Exporter CSV…",
     "Fichier `*_adobe.csv` avec BOM UTF-8, colonnes Filename,Title,Keywords,Category,Releases",
     "critique"),
    ("T-041", "Sorties", "REQ-OUT-02", "CSV Shutterstock — colonnes",
     "Export CSV Shutterstock après un rapport expert",
     "Rapport expert calculé", "Bouton Exporter CSV…",
     "Fichier `*_shutterstock.csv` colonnes Filename,Description,Keywords,Categories,Editorial,Mature,Illustration",
     "critique"),
    ("T-042", "Sorties", "REQ-OUT-03", "Keywords séparateur virgule (fix P0)",
     "Vérifier que Keywords est `kw1, kw2, kw3` (virgule), PAS `kw1 kw2 kw3` (espace)",
     "Au moins 3 keywords dans le rapport", "Ouvrir le CSV produit",
     "Cellule Keywords contient au moins une « , » entre chaque mot-clé",
     "critique"),
    ("T-043", "Sorties", "REQ-OUT-04", "Mapping catégorie SH→Adobe",
     "Input catégorie SH = « Business/Finance » → output Adobe Category = « Business »",
     "Image avec catégorie SH", "Export Adobe",
     "Colonne Category = « Business » dans le CSV Adobe",
     "majeure"),
    ("T-044", "Sorties", "REQ-OUT-05", "Export batch 50 lignes",
     "Sélection multiple de 50 images → 1 export",
     "50 images scannées et sélectionnées", "Export double CSV",
     "Chaque CSV contient exactement 50 lignes + 1 ligne d'en-tête",
     "majeure"),

    # --- E. Cas limites ---
    ("T-050", "Cas limites", "REQ-LIM-01", "0 keyword",
     "Image avec IPTC sans keywords", "input_no_keywords.jpg",
     "Rapport expert",
     "RejectionRisk severity=blocker « minimum 7 »", "majeure"),
    ("T-051", "Cas limites", "REQ-LIM-02", ">50 keywords",
     "Image avec 80 keywords dans IPTC", "input_many_keywords.jpg",
     "Rapport expert", "Tronqué à 50, pas d'erreur, warning éventuel", "mineure"),
    ("T-052", "Cas limites", "REQ-LIM-03", "Titre vide",
     "IPTC sans headline ni object_name", "input_no_title.jpg",
     "Rapport expert",
     "Fallback titre depuis nom de fichier, ou RejectionRisk « Titre absent »",
     "mineure"),
    ("T-053", "Cas limites", "REQ-LIM-04", "Marques filtrées",
     "Keywords contiennent `apple`, `nike`, `coca-cola`",
     "input_brands.jpg", "Rapport expert",
     "Aucune marque ne survit dans report.keywords",
     "majeure"),
    ("T-054", "Cas limites", "REQ-LIM-05", "Stuffing hors titre filtré",
     "Keywords contiennent `stock`, `image`, `wallpaper`, titre = « Sunset »",
     "input_stuffing.jpg", "Rapport expert",
     "Ces mots-clés filtrés silencieusement", "mineure"),
    ("T-055", "Cas limites", "REQ-LIM-06", "Stuffing dans titre conservé",
     "Keywords contiennent `photo`, titre = « Lake Photo »",
     "input_stuffing_in_title.jpg", "Rapport expert",
     "« photo » est conservé (présent dans le titre)", "mineure"),

    # --- F. Performance ---
    ("T-060", "Performance", "REQ-PERF-01", "Heuristique 50 images",
     "Construire 50 rapports sans IA", "50 images scannées",
     "Chronométrer batch heuristique",
     "P95 < 5 s, P99 < 8 s sur machine de référence", "mineure"),
    ("T-061", "Performance", "REQ-PERF-02", "Export 500 lignes",
     "Export double CSV de 500 rapports", "500 rapports calculés",
     "Chronométrer export_double_csv", "Total < 2 s", "mineure"),
    ("T-062", "Performance", "REQ-PERF-03", "Ouverture modal expert",
     "Latence d'ouverture modal Rapport expert", "Image sélectionnée",
     "Chronométrer entre clic et premier widget visible",
     "< 500 ms (rendu) + worker thread asynchrone", "mineure"),

    # --- G. Robustesse ---
    ("T-070", "Robustesse", "REQ-ROB-01", "Backend api=None",
     "Lancer l'app sans backend", "Démarrage forcé api=None",
     "Ouvrir Rapport expert + Validation",
     "EmptyState propre sur chaque modal, pas de crash", "majeure"),
    ("T-071", "Robustesse", "REQ-ROB-02", "ExifTool absent",
     "ExifTool désinstallé / chemin invalide", "PATH sans exiftool",
     "Tenter Lire/Écrire IPTC",
     "Toast warning « ExifTool absent », lecture/écriture désactivées proprement",
     "majeure"),
    ("T-072", "Robustesse", "REQ-ROB-03", "Ollama absent + Use_AI=True",
     "Use_AI coché alors qu'Ollama n'est pas installé",
     "Service Ollama stoppé", "Régénérer rapport en mode IA",
     "Fallback heuristique transparent, log warning, pas de crash",
     "majeure"),
    ("T-073", "Robustesse", "REQ-ROB-04", "Disque plein à l'export",
     "Tenter export CSV vers chemin sans espace",
     "Disque/dossier en lecture seule", "Export double CSV",
     "Erreur OS contenue, toast d'erreur clair, app stable", "mineure"),
    ("T-074", "Robustesse", "REQ-ROB-05", "Permission refusée écriture IPTC",
     "Fichier en lecture seule", "chmod 444",
     "Cliquer Écrire dans l'éditeur",
     "Erreur contenue, toast d'erreur, fichier inchangé", "mineure"),

    # --- H. Régression ---
    ("T-080", "Régression", "REQ-REG-01", "Mode Shutterstock legacy",
     "vision_analyzer.analyze_image() doit retourner le même schéma qu'avant la refonte",
     "Ollama présent + image", "Appeler l'API legacy directement",
     "AnalysisResult avec title/description/keywords/categories — non régressé",
     "critique"),
    ("T-081", "Régression", "REQ-REG-02", "Suite pytest existante",
     "120 tests doivent rester verts (90 v2.0 + 30 licensing v2.1)", "—", "pytest tests/ -q",
     "120 passed, 0 failed", "critique"),
    ("T-082", "Régression", "REQ-REG-03", "Build PyInstaller",
     "build.py debug + release sortent un EXE qui démarre",
     "—", "python build.py all",
     "2 EXE produits, smoke test mainloop OK, exit code 0",
     "critique"),
]


# ============================================================================

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUMMARY_FILL = PatternFill("solid", fgColor="E7F0F9")
SUMMARY_FONT = Font(bold=True, color="1F4E78", size=11)


def _make_tests_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Tests"

    # Header
    for col, label in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    # Rows
    for i, t in enumerate(TESTS, start=2):
        tid, cat, req, feat, desc, prereq, inp, expected, severity = t
        row = [tid, cat, req, feat, desc, prereq, inp, expected,
               "", "NA", severity, "", "", ""]
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER

    # Column widths
    widths = [10, 13, 14, 22, 38, 22, 28, 50, 28, 9, 11, 12, 12, 26]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    # Data validation: Statut + Sévérité
    nb = len(TESTS) + 1
    dv_statut = DataValidation(type="list", formula1='"OK,NOK,NA,Bloqué"', allow_blank=True)
    dv_statut.add(f"J2:J{nb}")
    ws.add_data_validation(dv_statut)

    dv_sev = DataValidation(type="list", formula1='"critique,majeure,mineure"', allow_blank=True)
    dv_sev.add(f"K2:K{nb}")
    ws.add_data_validation(dv_sev)

    # Conditional colours on Statut
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    gray = PatternFill("solid", fgColor="EAEAEA")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    ws.conditional_formatting.add(
        f"J2:J{nb}", CellIsRule(operator="equal", formula=['"OK"'], fill=green))
    ws.conditional_formatting.add(
        f"J2:J{nb}", CellIsRule(operator="equal", formula=['"NOK"'], fill=red))
    ws.conditional_formatting.add(
        f"J2:J{nb}", CellIsRule(operator="equal", formula=['"NA"'], fill=gray))
    ws.conditional_formatting.add(
        f"J2:J{nb}", CellIsRule(operator="equal", formula=['"Bloqué"'], fill=yellow))

    # Freeze header + ID column
    ws.freeze_panes = "B2"


def _make_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Synthèse", 0)

    # Title block
    ws["A1"] = "Synthèse de qualification"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
    ws.merge_cells("A1:F1")

    info = [
        ("Outil", "ShutterstockAnalyzer"),
        ("Version", "v2.1.0"),
        ("Date génération", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Nombre de tests", len(TESTS)),
        ("Environnement", "Windows 10/11 — Python 3.11"),
        ("Testeur principal", ""),
    ]
    for i, (k, v) in enumerate(info, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    # Categories table
    start = len(info) + 5
    ws.cell(row=start, column=1, value="Catégorie").font = HEADER_FONT
    ws.cell(row=start, column=1).fill = HEADER_FILL
    for col, label in enumerate(["Total", "OK", "NOK", "NA", "Bloqué", "Taux OK"], start=2):
        c = ws.cell(row=start, column=col, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    cats = sorted({t[1] for t in TESTS})
    for i, cat in enumerate(cats, start=start + 1):
        ws.cell(row=i, column=1, value=cat).font = SUMMARY_FONT
        ws.cell(row=i, column=2,
                value=f'=COUNTIF(Tests!B:B,"{cat}")')
        ws.cell(row=i, column=3,
                value=f'=COUNTIFS(Tests!B:B,"{cat}",Tests!J:J,"OK")')
        ws.cell(row=i, column=4,
                value=f'=COUNTIFS(Tests!B:B,"{cat}",Tests!J:J,"NOK")')
        ws.cell(row=i, column=5,
                value=f'=COUNTIFS(Tests!B:B,"{cat}",Tests!J:J,"NA")')
        ws.cell(row=i, column=6,
                value=f'=COUNTIFS(Tests!B:B,"{cat}",Tests!J:J,"Bloqué")')
        ws.cell(row=i, column=7,
                value=f'=IFERROR(C{i}/(B{i}-E{i}),0)')
        ws.cell(row=i, column=7).number_format = "0.0%"

    # Totals row
    tot = start + 1 + len(cats)
    ws.cell(row=tot, column=1, value="TOTAL").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=tot, column=1).fill = HEADER_FILL
    for col in range(2, 7):
        letter = get_column_letter(col)
        ws.cell(row=tot, column=col,
                value=f"=SUM({letter}{start+1}:{letter}{tot-1})")
        ws.cell(row=tot, column=col).font = Font(bold=True)
    ws.cell(row=tot, column=7,
            value=f"=IFERROR(C{tot}/(B{tot}-E{tot}),0)")
    ws.cell(row=tot, column=7).number_format = "0.0%"
    ws.cell(row=tot, column=7).font = Font(bold=True)

    # Conclusion area
    concl = tot + 3
    ws.cell(row=concl, column=1, value="Conclusion").font = Font(bold=True, size=12)
    for i, label in enumerate(
        ["Décision (go / no-go / go conditionnel)", "Signature", "Date"], start=concl + 1
    ):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)

    # Widths
    widths = [38, 10, 10, 10, 10, 10, 12]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def main() -> None:
    wb = Workbook()
    _make_tests_sheet(wb)
    _make_summary_sheet(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"OK -> {OUT}  ({len(TESTS)} tests)")


if __name__ == "__main__":
    main()
